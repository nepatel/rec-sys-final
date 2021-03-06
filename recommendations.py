'''
1. Create content-based recommenders (Feature Encoding, TF-IDF/CosineSim
       using item/genre feature data)
2. Create hybrid recommenders: TF-IDF (base recommender), augment CosineSim matrix
       using a pre-computed item-item similarity matrix
3. Implement LOOCV and write trial parameters and results to Excel

Programmer names: Joseph Brock, Jake Carver, Neil Patel, Annabel Winters-McCabe

Collaborator/Author: Carlos Seminario

sources:
https://www.freecodecamp.org/news/how-to-process-textual-data-using-tf-idf-in-python-cd2bbc0a94a3/
http://blog.christianperone.com/2013/09/machine-learning-cosine-similarity-for-vector-space-models-part-iii/
https://kavita-ganesan.com/tfidftransformer-tfidfvectorizer-usage-differences/#.XoT9p257k1L

reference:
https://scikit-learn.org/stable/modules/generated/sklearn.feature_extraction.text.CountVectorizer.html

'''
#################
##  LIBRARIES  ##
#################
import numpy as np
import pandas as pd
import math
import os
import pickle
from copy import deepcopy
from math import sqrt
from matplotlib import pyplot as plt
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.model_selection import LeaveOneOut
from sklearn.metrics import mean_squared_error
from sklearn.metrics import mean_absolute_error
from openpyxl import load_workbook


#################
##  CONSTANTS  ##
#################
# accept all positive similarities > [i] for TF-IDF/ConsineSim Recommender
SIM_THRESHOLDS = [0, 0.3, 0.5, 0.7]

# we went with n/25 b/c of good mixture of accuracy and coverage from midterm results
SIM_WEIGHTING = 25

# we went with 0.75 b/c of best accuracy metrics with LOOCV on Euclidean/Pearson
HYBRID_WEIGHTING = 0.75


####################
##  HELPER FUNCS  ##
####################
def from_file_to_2D(path, genrefile, itemfile):
    '''
    Load feature matrix from specified file

    Parameters:
        -- path: directory path to datafile and itemfile
        -- genrefile: delimited file that maps genre to genre index
        -- itemfile: delimited file that maps itemid to item name and genre

    Returns:
        -- movies: a dictionary containing movie titles (value) for a given movieID (key)
        -- genres: dictionary, key is genre, value is index into row of features array
        -- features: a 2D list of features by item, values are 1 and 0;
                     rows map to items and columns map to genre
                     returns as np.array()
    '''

    # Get movie titles, place into movies dictionary indexed by itemID
    movies = {}
    try:
        with open(path + '/' + itemfile, encoding='iso8859') as myfile:
            # this encoding is required for some datasets: encoding='iso8859'
            for line in myfile:
                (id, title) = line.split('|')[0:2]
                movies[id] = title.strip()

    # Error processing
    except UnicodeDecodeError as ex:
        print(ex)
        print(len(movies), line, id, title)
        return {}
    except ValueError as ex:
        print('ValueError', ex)
        print(len(movies), line, id, title)
    except Exception as ex:
        print(ex)
        print(len(movies))
        return {}

    # Get movie genre from the genre file, place into genre dictionary indexed by genre index
    genres = {}  # key is genre index, value is the genre string

    # construct generes dictionary
    try:
        with open(path + '/' + genrefile, encoding='iso8859') as myotherfile:
            # this encoding is required for some datasets: encoding='iso8859'
            id = 0
            for line in myotherfile:
                genre = line.split('|')[0].strip()
                if genre == '':
                    continue
                genres[id] = genre
                id += 1

    # Error processing
    except UnicodeDecodeError as ex:
        print(ex)
        print(len(movies), line, id, title)
        return {}
    except ValueError as ex:
        print('ValueError', ex)
        print(len(movies), line, id, title)
    except Exception as ex:
        print(ex)
        print(len(movies))
        return {}

    print(genres)

    # Load data into a nested 2D list
    features = []
    start_feature_index = 5
    try:
        for line in open(path+'/' + itemfile, encoding='iso8859'):
            # print(line, line.split('|')) #debug
            fields = line.split('|')[start_feature_index:]
            row = []
            for feature in fields:
                row.append(int(feature))
            features.append(row)
        features = np.array(features)
    except Exception as ex:
        print(ex)
        print('Proceeding with len(features)', len(features))
        # return {}
    # return features matrix
    return movies, genres, features


def from_file_to_dict(path, datafile, itemfile):
    '''
    Loads user-item matrix from specified file

    Parameters:
        -- path: directory path to datafile and itemfile
        -- datafile: delimited file containing userid, itemid, rating
        -- itemfile: delimited file that maps itemid to item name

    Returns:
        -- prefs: a nested dictionary containing item ratings (value) for each user (key)
    '''

    # Get movie titles, place into movies dictionary indexed by itemID
    movies = {}
    try:
        with open(path + '/' + itemfile, encoding='iso8859') as myfile:
            # this encoding is required for some datasets: encoding='iso8859'
            for line in myfile:
                (id, title) = line.split('|')[0:2]
                movies[id] = title.strip()

    # Error processing
    except UnicodeDecodeError as ex:
        print(ex)
        print(len(movies), line, id, title)
        return {}
    except ValueError as ex:
        print('ValueError', ex)
        print(len(movies), line, id, title)
    except Exception as ex:
        print(ex)
        print(len(movies))
        return {}

    # Load data into a nested dictionary
    prefs = {}
    for line in open(path+'/' + datafile):
        # print(line, line.split('\t')) #debug
        (user, movieid, rating, ts) = line.split('\t')
        user = user.strip()  # remove spaces
        movieid = movieid.strip()  # remove spaces
        prefs.setdefault(user, {})  # make it a nested dicitonary
        prefs[user][movies[movieid]] = float(rating)

    # return a dictionary of preferences
    return prefs


def transformPrefs(prefs):
    '''
    Transposes U-I matrix (prefs dictionary)
    
    Parameters:
        -- prefs: dictionary containing user-item matrix
    
    Returns:
        -- A transposed U-I matrix, i.e., if prefs was a U-I matrix,
           this function returns an I-U matrix
    '''

    result = {}
    for person in prefs:
        for item in prefs[person]:
            result.setdefault(item, {})
            result[item][person] = prefs[person][item]   # Flip item and person
    return result


def prefs_to_2D_list(prefs):
    '''
    Converts prefs dictionary into 2D list used as input for the MF class

    Parameters:
        -- prefs: user-item matrix as a dicitonary (dictionary)

    Returns:
        -- ui_matrix: (list) contains user-item matrix as a 2D list
    '''

    ui_matrix = []

    user_keys_list = list(prefs.keys())
    num_users = len(user_keys_list)
    # print (len(user_keys_list), user_keys_list[:10]) # debug

    itemPrefs = transformPrefs(prefs)  # traspose the prefs u-i matrix
    item_keys_list = list(itemPrefs.keys())
    num_items = len(item_keys_list)
    # print (len(item_keys_list), item_keys_list[:10]) # debug

    sorted_list = True  # <== set manually to test how this affects results

    if sorted_list == True:
        user_keys_list.sort()
        item_keys_list.sort()
        print('\nsorted_list =', sorted_list)

    # initialize a 2D matrix as a list of zeroes with
    #     num users (height) and num items (width)

    for i in range(num_users):
        row = []
        for j in range(num_items):
            row.append(0.0)
        ui_matrix.append(row)

    # populate 2D list from prefs
    # Load data into a nested list

    for user in prefs:
        for item in prefs[user]:
            user_idx = user_keys_list.index(user)
            movieid_idx = item_keys_list.index(item)

            try:
                # make it a nested list
                ui_matrix[user_idx][movieid_idx] = prefs[user][item]
            except Exception as ex:
                print(ex)
                print(user_idx, movieid_idx)

    # return 2D user-item matrix
    return ui_matrix


def to_array(prefs):
    ''' Converts prefs dictionary into 2D list '''

    R = prefs_to_2D_list(prefs)
    R = np.array(R)
    print('to_array -- height: %d, width: %d' % (len(R), len(R[0])))
    return R


def to_string(features):
    ''' Converts features np.array into list of feature strings '''

    feature_str = []
    for i in range(len(features)):
        row = ''
        for j in range(len(features[0])):
            row += (str(features[i][j]))
        feature_str.append(row)
    print('to_string -- height: %d, width: %d' %
          (len(feature_str), len(feature_str[0])))
    return feature_str


def to_docs(features_str, genres):
    ''' Converts feature strings to a list of doc strings for TFIDF '''

    feature_docs = []
    for doc_str in features_str:
        row = ''
        for i in range(len(doc_str)):
            if doc_str[i] == '1':
                # map the indices to the actual genre string
                row += (genres[i] + ' ')
        # and remove that pesky space at the end
        feature_docs.append(row.strip())

    print('to_docs -- height: %d, width: varies' % (len(feature_docs)))
    return feature_docs


def movie_to_ID(movies):
    ''' Converts movies mapping from "id to title" to "title to id" '''

    movie_title_to_id = {}
    for id in movies.keys():
        movie_title_to_id[movies[id]] = id
    return movie_title_to_id


########################
##  SIMILARITY FUNCS  ##  
########################
def cosine_sim(docs):
    '''
    Performs cosine sim calcs on features list, aka docs in TF-IDF world

    Parameters:
        -- docs: list of item features

    Returns:
        -- list containing cosim_matrix: item_feature-item_feature cosine similarity matrix
    '''

    print()
    print('## Cosine Similarity calc ##')
    print()
    print('Documents:', docs[:10])

    print()
    print('## Count and Transform ##')
    print()

    # choose one of these invocations
    tfidf_vectorizer = TfidfVectorizer()  # orig

    tfidf_matrix = tfidf_vectorizer.fit_transform(docs)
    # print (tfidf_matrix.shape, type(tfidf_matrix)) # debug

    print()
    print('Document similarity matrix:')
    cosim_matrix = cosine_similarity(tfidf_matrix[0:], tfidf_matrix)
    print(type(cosim_matrix), len(cosim_matrix))
    print()
    print(cosim_matrix[0:6])
    print()

    return cosim_matrix


def sim_distance(prefs, p1, p2, sim_weighting=0):
    '''
    Calculate Euclidean distance similarity

    Parameters:
        -- prefs: dictionary containing user-item matrix
        -- p1: string containing name of user 1
        -- p2: string containing name of user 2
        -- sim_weighting: similarity significance weighting factor (0, 25, 50)
                          [default is 0, which represents No Weighting]

    Returns:
        -- Euclidean distance similarity as a float
    '''

    # Get the list of shared_items
    si = {}
    for item in prefs[p1]:
        if item in prefs[p2]:
            si[item] = 1

    # if they have no ratings in common, return 0
    if len(si) == 0:
        return 0

    # Add up the squares of all the differences
    sum_of_squares = sum([pow(prefs[p1][item]-prefs[p2][item], 2)
                          for item in prefs[p1] if item in prefs[p2]])

    distance_sim = 1/(1+sqrt(sum_of_squares))

    # apply significance weighting, if any
    if sim_weighting != 0:
        if len(si) < sim_weighting:
            distance_sim *= (len(si) / sim_weighting)

    return distance_sim


def sim_pearson(prefs, p1, p2, sim_weighting=0):
    '''
    Calculate Pearson Correlation similarity

    Parameters:
        -- prefs: dictionary containing user-item matrix
        -- p1: string containing name of user 1
        -- p2: string containing name of user 2
        -- sim_weighting: similarity significance weighting factor (0, 25, 50)
                          [default is 0, which represents No Weighting]

    Returns:
        -- Pearson Correlation similarity as a float
    '''

    # Get the list of shared_items
    si = {}
    for item in prefs[p1]:
        if item in prefs[p2]:
            si[item] = 1

    # if they have no ratings in common, return 0
    if len(si) == 0:
        return 0

    # calc avg rating for p1 and p2, using only shared ratings
    x_avg = 0
    y_avg = 0

    for item in si:
        x_avg += prefs[p1][item]
        y_avg += prefs[p2][item]

    x_avg /= len(si)
    y_avg /= len(si)

    # calc numerator of Pearson correlation formula
    numerator = sum([(prefs[p1][item] - x_avg) * (prefs[p2][item] - y_avg)
                     for item in si])

    # calc denominator of Pearson correlation formula
    denominator = math.sqrt(sum([(prefs[p1][item] - x_avg)**2 for item in si])) * \
        math.sqrt(sum([(prefs[p2][item] - y_avg)**2 for item in si]))

    # catch divide-by-0 errors
    if denominator != 0:
        sim_pearson = numerator / denominator

        # apply significance weighting, if any
        if sim_weighting != 0:
            sim_pearson *= (len(si) / sim_weighting)

        return sim_pearson
    else:
        return 0


##########################
##  TOP-N RECOMMENDERS  ##
##########################
def get_TFIDF_recommendations(prefs, cosim_matrix, user, sim_threshold, movies, n=15):
    '''
    Calculates recommendations for a given user

    Parameters:
        -- prefs: dictionary containing user-item matrix
        -- cosim_matrix: list containing item_feature-item_feature cosine similarity matrix
        -- user: string containing name of user requesting recommendation
        -- sim_threshold: float that determines the minimum similarity to be a "neighbor"
        -- movies: dictionary that maps movieid to movie title
        -- n: number of recommendations to print

    Returns:
        -- rankings: A list of recommended items with 0 or more tuples,
           each tuple contains (predicted rating, item name).
           List is sorted, high to low, by predicted rating.
           An empty list is returned when no recommendations have been calc'd.
    '''

    recs = []
    userRatings = prefs[str(user)]

    # iterate through cosim_matrix
    for i in range(1, len(movies)+1):
        # movie is already rated by user
        if movies[str(i)] in userRatings:
            continue

        num = 0
        denom = 0

        for j in range(1, len(movies) + 1):
            # neighbor movie has not been rated by the user
            if movies[str(j)] not in userRatings:
                continue

            # do not compare self to self
            if i == j:
                continue

            # does not meet similarity threshold
            if cosim_matrix[i-1][j-1] < sim_threshold:
                continue

            num += userRatings[movies[str(j)]]*cosim_matrix[i-1][j-1]
            denom += cosim_matrix[i-1][j-1]

        if num > 0 and denom > 0:
            recs.append((num/denom, movies[str(i)]))

    print(sorted(recs, reverse=True)[:n])
    return recs


def get_FE_recommendations(prefs, features, movie_title_to_id, movies, user, n=15):
    '''
    Calculates recommendations for a given user

    Parameters:
        -- prefs: dictionary containing user-item matrix
        -- features: np.array whose height is based on number of items
                     and width equals the number of unique features (e.g., genre)
        -- movie_title_to_id: dictionary that maps movie title to movieid
        -- movies: dictionary that maps movieid to movie title
        -- user: string containing name of user requesting recommendation
        -- n: number of recommendations to return if greater than 10

    Returns:
        -- rankings: A list of recommended items with 0 or more tuples,
           each tuple contains (predicted rating, item name).
           List is sorted, high to low, by predicted rating.
           An empty list is returned when no recommendations have been calc'd.

    '''

    # generate set of total possible ids
    total_ids = list(range(0, len(features)))
    total_set = set(total_ids)

    feature_preference = np.copy(features)

    # transform features into the feature_preference matrix
    feature_preference = feature_preference.astype('float64')
    rated = set()
    for movie in prefs[user]:
        id = (int)(movie_title_to_id[movie])-1
        feature_preference[id] *= prefs[user][movie]
        rated.add(id)

    # set subtraction to pull out unrated items
    unrated_ids = total_set.difference(rated)

    # set unrated rows to 0s
    for id in unrated_ids:
        feature_preference[id] *= 0

    # take column wise sum, overall sum, and normalized vector
    col_sums = np.sum(feature_preference, axis=0)
    overall_sum = np.sum(feature_preference)
    norm_vector = col_sums/overall_sum

    recs = []

    # for each unrated item
    for id in unrated_ids:
        # multiply features row for item by normalized vector
        norm_weight = features[id]*norm_vector
        norm_sum = np.sum(norm_weight)

        # avoid divide by 0 error
        if norm_sum == 0:
            continue
        norm_weight = norm_weight/norm_sum

        # get nonzero count
        nonzero_count = np.count_nonzero(feature_preference, axis=0)

        # get vector of averages, pass over divide by 0
        avgs = col_sums/nonzero_count

        # remove irrelevant features
        avgs *= features[id].astype('float64')
        weight_avg = avgs*norm_weight
        final_rec = np.nansum(weight_avg)
        recs.append((final_rec, movies[str(id+1)]))

        # sort high to low
        recs = sorted(recs, reverse=True)

        if len(recs) > 10:
            recs = recs[:n]

    return recs


def get_hybrid_recommendations(prefs, cosim_matrix, user, sim_threshold, movies, ii_matrix, weighted=False, n=15):
    '''
    Calculates recommendations for a given user

    Parameters:
        -- prefs: dictionary containing user-item matrix
        -- cosim_matrix: list containing item_feature-item_feature cosine similarity matrix
        -- user: string containing name of user requesting recommendation
        -- sim_threshold: float that determines the minimum similarity to be a "neighbor"
        -- movies: dictionary mapping movieid to movie title
        -- ii_matrix: pre-computed item-item matrix from a collaborative filtering
        -- weighted: if true, replace sim in cosim_matrix with ii_value times HYBRID_WEIGHT
        -- n: number of recommendations to print

    Returns:
        -- rankings: A list of recommended items with 0 or more tuples,
           each tuple contains (predicted rating, item name).
           List is sorted, high to low, by predicted rating.
           An empty list is returned when no recommendations have been calc'd.
    '''

    recs = []
    userRatings = prefs[str(user)]
    copy_cosim = deepcopy(cosim_matrix)

    # iterate through cosim_matrix
    for i in range(1, len(copy_cosim) + 1):
        # movie is already rated by user
        if movies[str(i)] in userRatings:
            continue

        num = 0
        denom = 0

        for j in range(1, len(copy_cosim) + 1):
            # neighbor movie has not been rated by the user
            if movies[str(j)] not in userRatings:
                continue

            # do not compare self to self
            if i == j:
                continue

            # does not meet similarity threshold
            if copy_cosim[i-1][j-1] < sim_threshold:
                continue

            # "fill in" similarities of 0 using collaborative filtering item-item similarity matrix
            if copy_cosim[i-1][j-1] == 0:

                # skip missing titles because there is no similarity
                if movies[str(j)] not in ii_matrix[movies[str(i)]]:
                    continue
                elif ii_matrix[movies[str(i)]][movies[str(j)]] <= 0:
                    continue
                # replace with corresponding item-item similarity matrix value
                copy_cosim[i-1][j-1] = ii_matrix[movies[str(i)]][movies[str(j)]]

                # compute similarity with weighting factor
                if weighted:
                    copy_cosim[i-1][j-1] *= HYBRID_WEIGHTING

            num += userRatings[movies[str(j)]] * copy_cosim[i-1][j-1]
            denom += copy_cosim[i-1][j-1]

        if num > 0 and denom > 0:
            recs.append((num/denom, movies[str(i)]))

    print(sorted(recs, reverse=True)[:n])
    return recs


def topMatches(prefs, person, n=100, similarity=sim_pearson, sim_weighting=0, sim_threshold=0):
    '''
    Returns the best matches for person from the prefs dictionary

    Parameters:
        -- prefs: dictionary containing user-item matrix
        -- person: string containing name of user
        -- n: number of recommendations to return
        -- similarity: function to calc similarity [sim_pearson is default]
        -- sim_weighting: similarity significance weighting factor (0, 25, 50)
                          [default is 0, which represents No Weighting]
        -- sim_threshold: float that determines the minimum similarity to be a "neighbor"

    Returns:
        -- A list of similar matches with 0 or more tuples,
           each tuple contains (similarity, item name).
           List is sorted, high to low, by similarity.
           An empty list is returned when no matches have been calc'd.
    '''

    scores = []
    scores_dict = {}

    # iterate through users in prefs
    for other in prefs:
        # calculate similarity score
        score = similarity(prefs, person, other, sim_weighting)

        # don't compare me to myself, accept scores above the threshold
        if other != person:
            scores.append((score, other))

    scores = sorted(scores, reverse=True)
    scores = scores[:n]

    for i in range(len(scores)):
        scores_dict[scores[i][1]] = scores[i][0]

    return scores_dict


def calculateSimilarItems(prefs, similarity=sim_pearson, sim_weighting=SIM_WEIGHTING, sim_threshold=0):
    '''
    Creates a dictionary of items showing which other items they are most similar to.

    Parameters:
        -- prefs: dictionary containing user-item matrix
        -- similarity: function to calc similarity (sim_pearson is default)
        -- sim_weighting: similarity significance weighting factor (0, 25, 50)
                          [default is 0, which represents No Weighting]
        -- sim_threshold: float that determines the minimum similarity to be a "neighbor"

    Returns:
        -- A dictionary with a similarity matrix
    '''

    result = {}
    c = 0

    # Invert the preference matrix to be item-centric
    itemPrefs = transformPrefs(prefs)

    for item in itemPrefs:
      # Status updates for larger datasets
        c += 1
        if c % 100 == 0:
            percent_complete = (100*c)/len(itemPrefs)
            print("%d%% complete" % (percent_complete))

        # Find the most similar items to this one
        scores = topMatches(itemPrefs, item, 100, similarity,
                            sim_weighting, sim_threshold)
        result[item] = scores

    return result


###########################
##  SINGLE RECOMMENDERS  ##
###########################
def get_TFIDF_recommendations_single(prefs, cosim_matrix, user, sim_threshold, movies, movies2, ii_matrix, excluded):
    '''
    Calculates recommendations for a given user

    Parameters:
        -- prefs: dictionary containing user-item matrix
        -- cosim_matrix: list containing item_feature-item_feature cosine similarity matrix
        -- user: string containing name of user requesting recommendation
        -- sim_threshold: float that determines the minimum similarity to be a "neighbor"
        -- movies: dictionary mapping movieid to movie title
        -- movies2: dictionary that maps movie title to movieid
        -- ii_matrix: pre-computed item-item matrix from a collaborative filtering
        -- excluded: string of the item "left out" by LOOCV

    Returns:
        -- rankings: A list of recommended items with 0 or more tuples,
           each tuple contains (predicted rating, item name).
           List is sorted, high to low, by predicted rating.
           An empty list is returned when no recommendations have been calc'd.
    '''

    userRatings = prefs[str(user)]

    num = 0
    denom = 0

    # iterate through cosim_matrix
    for j in range(1, len(movies) + 1):
        # movie is already rated by user
        if movies[str(j)] not in userRatings:
            continue

         # do not compare self to self
        if int(movies2[excluded]) == j:
            continue

        # does not meet similarity threshold
        if cosim_matrix[int(movies2[excluded])-1][j-1] < sim_threshold:
            continue

        num += userRatings[movies[str(j)]] * \
            cosim_matrix[int(movies2[excluded])-1][j-1]
        denom += cosim_matrix[int(movies2[excluded])-1][j-1]

    if num > 0 and denom > 0:
        return (num/denom, excluded)

    return (None, excluded)


def get_FE_recommendations_single(prefs, features, user, sim_threshold, movies, movies2, ii_matrix, excluded):
    '''
    Calculates recommendations for a given user

    Parameters:
        -- prefs: dictionary containing user-item matrix
        -- features: an np.array whose height is based on number of items
                     and width equals the number of unique features (e.g., genre)
        -- user: string containing name of user requesting recommendation
        -- sim_threshold: float that determines the minimum similarity to be a "neighbor"
        -- movies: dictionary mapping movieid to movie title
        -- movies2: dictionary that maps movie title to movieid
        -- ii_matrix: pre-computed item-item matrix from a collaborative filtering
        -- excluded: string of the item "left out" by LOOCV

    Returns:
        -- rankings: A list of recommended items with 0 or more tuples,
           each tuple contains (predicted rating, item name).
           List is sorted, high to low, by predicted rating.
           An empty list is returned when no recommendations have been calc'd.

    '''

    # generate set of total possible ids
    total_ids = list(range(0, len(features)))
    total_set = set(total_ids)

    feature_preference = np.copy(features)

    # transform features into the feature_preference matrix
    feature_preference = feature_preference.astype('float64')
    rated = set()
    for movie in prefs[user]:
        id = (int)(movies2[movie])-1
        feature_preference[id] *= prefs[user][movie]
        rated.add(id)

    # set subtraction to pull out unrated items
    unrated_ids = total_set.difference(rated)

    # set unrated rows to 0s
    for id in unrated_ids:
        feature_preference[id] *= 0

    # take column wise sum, overall sum, and normalized vector
    col_sums = np.sum(feature_preference, axis=0)
    overall_sum = np.sum(feature_preference)
    norm_vector = col_sums/overall_sum

    rec = None

    # multiply features row for item by normalized vector
    norm_weight = features[int(movies2[excluded])-1]*norm_vector
    norm_sum = np.sum(norm_weight)

    if norm_sum == 0:
        return (None, excluded)
    try:
        norm_weight = norm_weight/norm_sum

        # get nonzero count
        nonzero_count = np.count_nonzero(feature_preference, axis=0)

        # get vector of averages, pass over divide by 0
        avgs = col_sums/nonzero_count

        # remove irrelevant features
        avgs *= features[int(movies2[excluded])-1].astype('float64')
        weight_avg = avgs*norm_weight
        final_rec = np.nansum(weight_avg)
        rec = (final_rec, excluded)

    except:
        return (None, excluded)

    return rec


def get_hybrid_recommendations_single(prefs, cosim_matrix, user, sim_threshold, movies, movies2, ii_matrix, excluded, weighted=True):
    '''
    Calculates recommendations for a given user

    Parameters:
        -- prefs: dictionary containing user-item matrix
        -- cosim_matrix: list containing item_feature-item_feature cosine similarity matrix
        -- user: string containing name of user requesting recommendation
        -- sim_threshold: float that determines the minimum similarity to be a "neighbor"
        -- movies: dictionary mapping movieid to movie title
        -- ii_matrix: pre-computed item-item matrix from a collaborative filtering
        -- excluded: string of the item "left out" by LOOCV
        -- weighted: if true, replace sim in cosim_matrix with ii_value times HYBRID_WEIGHT

    Returns:
        -- rankings: A list of recommended items with 0 or more tuples,
           each tuple contains (predicted rating, item name).
           List is sorted, high to low, by predicted rating.
           An empty list is returned when no recommendations have been calc'd.
    '''

    userRatings = prefs[str(user)]
    copy_cosim = deepcopy(cosim_matrix)
    
    num = 0
    denom = 0
    
    # iterate through cosim_matrix
    for j in range(1, len(copy_cosim) + 1):
        # neighbor movie has not been rated by the user
        if movies[str(j)] not in userRatings:
            continue

        # do not compare self to self
        if int(movies2[excluded]) == j:
            continue

        # does not meet similarity threshold
        if copy_cosim[int(movies2[excluded])-1][j-1] < sim_threshold:
            continue

        # "fill in" similarities of 0 using collaborative filtering item-item similarity matrix
        if copy_cosim[int(movies2[excluded])-1][j-1] == 0:

            # skip missing titles because there is no similarity
            if movies[str(j)] not in ii_matrix[excluded]:
                continue
            elif ii_matrix[excluded][movies[str(j)]] <= 0:
                continue

            # replace with corresponding item-item similarity matrix value
            copy_cosim[int(movies2[excluded])-1][j-1] = ii_matrix[excluded][movies[str(j)]]

            # compute similarity with weighting factor
            if weighted:
                copy_cosim[int(movies2[excluded])-1][j-1] *= HYBRID_WEIGHTING

        num += userRatings[movies[str(j)]] * copy_cosim[int(movies2[excluded])-1][j-1]
        denom += copy_cosim[int(movies2[excluded])-1][j-1]

    if num > 0 and denom > 0:
        return (num/denom, excluded)

    return (None, excluded)


######################################
##  LEAVE-ONE-OUT-CROSS-VALIDATION  ## 
######################################
def loo_cv_sim(prefs, sim, algo, sim_matrix, itemsim, movies, sim_threshold, ws, r, weighted=False):
    """
    Leave-One-Out-Cross-Evaluation: evaluates recommender system ACCURACY

    Parameters:
        -- prefs dataset: critics, etc.sim
        -- sim: distance, pearson, etc.
        -- algo: user-based recommender, item-based recommender, etc.
        -- sim_matrix: pre-computed similarity matrix
        -- itemsim: 
        -- movies: dictionary mapping movieid to movie title
        -- sim_threshold: float that determines the minimum similarity to be a "neighbor"
        -- ws: Excel worksheet object
        -- r: row index in Excel worksheet
        -- weighted: if true, replace sim in cosim_matrix with ii_value times HYBRID_WEIGHT

    Returns:
        -- error_total: MSE, or MAE, or RMSE totals for this set of conditions
        -- error_list: list of actual-predicted differences
    """

    true_list = []
    pred_list = []
    error_list = []
    count = 0

    # make a deep copy of prefs
    newPrefs = deepcopy(prefs)

    # swap movie ids and titles
    newMovies = {v: k for k, v in movies.items()}

    for i in list(prefs.keys()):
        checker = False

        for out in list(prefs[i].keys()):
            sumProd = 0
            sumSim = 0

            # make a copy of item, we will delete
            save = newPrefs[i][out]
            del newPrefs[i][out]

            # get recs for this item
            if weighted:
                rec = algo(newPrefs, sim_matrix, i, sim_threshold,
                           movies, newMovies, itemsim, out, weighted=weighted)
            else:
                rec = algo(newPrefs, sim_matrix, i, sim_threshold,
                           movies, newMovies, itemsim, out)
            newPrefs[i][out] = save

            if rec[1] == out and rec[0] != None and rec[0] != 0:
                try:
                    error = (prefs[i][rec[1]]-rec[0])**2
                    error_list.append(error)
                    true_list.append(prefs[i][rec[1]])
                    pred_list.append(rec[0])
                    checker = True
                except:
                    continue

        if len(prefs) < 20 or count % 50 == 0:
            print("User Num: ", count)
        count += 1

    # print results of LOOCV
    print('MSE: ', mean_squared_error(true_list, pred_list))
    print('MAE: ', mean_absolute_error(true_list, pred_list))
    print("RMSE: ", mean_squared_error(true_list, pred_list, squared=False))
    print("Coverage: ", len(error_list))
    print("Coverage PCT: ", len(error_list)/100000)

    # pipeline results of LOOCV to Excel spreadsheet
    # COMMENT OUT LINES 1018-1025 TO STOP EXCEL EXPORTS
    ws["A" + str(r)].value = str(algo)[14:-15]
    ws["B" + str(r)].value = sim
    ws["C" + str(r)].value = sim_threshold
    ws["D" + str(r)].value = weighted
    ws["E" + str(r)].value = mean_squared_error(true_list, pred_list)
    ws["F" + str(r)].value = mean_squared_error(true_list, pred_list, squared=False)
    ws["G" + str(r)].value = mean_absolute_error(true_list, pred_list)
    ws["H" + str(r)].value = len(error_list)/100000

    return error_list


def main():
    # Load critics dict from file
    path = os.getcwd()  # this gets the current working directory
    # you can customize path for your own computer here
    # print('\npath: %s' % path)  # debug
    prefs = {}
    done = False
    cosim_matrix = []
    itemsim = {}
    # open excel file in case writing from LCVSIM
    dest = "CSC_381_ALS_Results.xlsx"
    wb = load_workbook(filename=dest)
    ws = wb.create_sheet('results')
    row = 1
    while not done:
        print()
        file_io = input('R(ead) critics data from file?, \n'
                        'RML(ead) ml100K data from file?, \n'
                        'FE(ature Encoding) Setup?, \n'
                        'TFIDF(and cosine sim Setup)?, \n'
                        'CBR-FE(content-based recommendation Feature Encoding)?, \n'
                        'CBR-TF(content-based recommendation TF-IDF/CosineSim)? \n'
                        'SIM(ilarity matrix item-item calculation)? \n'
                        'HBR(content-based recommendation Hybrid)? \n'
                        'LCVSIM(leave-one-out-cross-validation)? \n'
                        '==>> '
                        )

        if file_io == 'R' or file_io == 'r':
            print()
            file_dir = 'data/'
            # for userids use 'critics_ratings_userIDs.data'
            datafile = 'critics_ratings.data'
            itemfile = 'critics_movies.item'
            genrefile = 'critics_movies.genre'  # movie genre file
            print('Reading "%s" dictionary from file' % datafile)
            prefs = from_file_to_dict(
                path, file_dir+datafile, file_dir+itemfile)
            movies, genres, features = from_file_to_2D(
                path, file_dir+genrefile, file_dir+itemfile)
            print('Number of users: %d\nList of users:' % len(prefs),
                  list(prefs.keys()))

            print('Number of distinct genres: %d, number of feature profiles: %d' % (
                len(genres), len(features)))
            print('genres')
            print(genres)
            print('features')
            print(features)

        elif file_io == 'RML' or file_io == 'rml':
            print()
            file_dir = 'data/ml-100k/'  # path from current directory
            datafile = 'u.data'  # ratngs file
            itemfile = 'u.item'  # movie titles file
            genrefile = 'u.genre'  # movie genre file
            print('Reading "%s" dictionary from file' % datafile)
            prefs = from_file_to_dict(
                path, file_dir+datafile, file_dir+itemfile)
            movies, genres, features = from_file_to_2D(
                path, file_dir+genrefile, file_dir+itemfile)

            print('Number of users: %d\nList of users [0:10]:'
                  % len(prefs), list(prefs.keys())[0:10])
            print('Number of distinct genres: %d, number of feature profiles: %d'
                  % (len(genres), len(features)))
            print('genres')
            print(genres)
            print('features')
            print(features)
            print(features.shape)

        elif file_io == 'FE' or file_io == 'fe':
            print()
            # movie_title_to_id = movie_to_ID(movies)
            # determine the U-I matrix to use ..
            if len(prefs) > 0 and len(prefs) <= 10:  # critics
                # convert prefs dictionary into 2D list
                R = to_array(prefs)

                '''
                # e.g., critics data (CES)
                R = np.array([
                [2.5, 3.5, 3.0, 3.5, 2.5, 3.0],
                [3.0, 3.5, 1.5, 5.0, 3.5, 3.0],
                [2.5, 3.0, 0.0, 3.5, 0.0, 4.0],
                [0.0, 3.5, 3.0, 4.0, 2.5, 4.5],
                [3.0, 4.0, 2.0, 3.0, 2.0, 3.0],
                [3.0, 4.0, 0.0, 5.0, 3.5, 3.0],
                [0.0, 4.5, 0.0, 4.0, 1.0, 0.0],
                ])
                '''
                print('critics')
                print(R)
                print()
                print('features')
                print(features)
                movie_title_to_id = movie_to_ID(movies)
                print(movie_title_to_id)
                for user in prefs:
                    print(user)
                    recs = get_FE_recommendations(
                        prefs, features, movie_title_to_id, movies, user)
                    print(user, recs)

            elif len(prefs) > 10:
                print('ml-100k')
                # convert prefs dictionary into 2D list
                R = to_array(prefs)
                movie_title_to_id = movie_to_ID(movies)
                recs = get_FE_recommendations(
                    prefs, features, movie_title_to_id, movies, '340')
                print(recs)
            else:
                print('Empty dictionary, read in some data!')
                print()

        elif file_io == 'TFIDF' or file_io == 'tfidf':
            print()
            # determine the U-I matrix to use ..
            if len(prefs) > 0 and len(prefs) <= 10:  # critics
                # convert prefs dictionary into 2D list
                R = to_array(prefs)
                feature_str = to_string(features)
                feature_docs = to_docs(feature_str, genres)

                '''
                # e.g., critics data (CES)
                R = np.array([
                [2.5, 3.5, 3.0, 3.5, 2.5, 3.0],
                [3.0, 3.5, 1.5, 5.0, 3.5, 3.0],
                [2.5, 3.0, 0.0, 3.5, 0.0, 4.0],
                [0.0, 3.5, 3.0, 4.0, 2.5, 4.5],
                [3.0, 4.0, 2.0, 3.0, 2.0, 3.0],
                [3.0, 4.0, 0.0, 5.0, 3.5, 3.0],
                [0.0, 4.5, 0.0, 4.0, 1.0, 0.0],
                ])
                '''
                print('critics')
                print(R)
                print()
                print('features')
                print(features)
                print()
                print('feature docs')
                print(feature_docs)
                cosim_matrix = cosine_sim(feature_docs)
                print()
                print('cosine sim matrix')
                print(cosim_matrix)

                graphArray = []
                for i in range(len(cosim_matrix)):
                    for j in range(i):
                        if cosim_matrix[i][j] != 0 and cosim_matrix[i][j] != 1:
                            graphArray.append(cosim_matrix[i][j])
                '''
                <class 'numpy.ndarray'>

                [[1.         0.         0.35053494 0.         0.         0.61834884]
                [0.         1.         0.19989455 0.17522576 0.25156892 0.        ]
                [0.35053494 0.19989455 1.         0.         0.79459157 0.        ]
                [0.         0.17522576 0.         1.         0.         0.        ]
                [0.         0.25156892 0.79459157 0.         1.         0.        ]
                [0.61834884 0.         0.         0.         0.         1.        ]]
                '''

                # print and plot histogram of similarites
                # plt.hist(graphArray, 10)
                # plt.show()

            elif len(prefs) > 10:
                print('ml-100k')
                # convert prefs dictionary into 2D list
                R = to_array(prefs)
                feature_str = to_string(features)
                feature_docs = to_docs(feature_str, genres)

                print(R[:3][:5])
                print()
                print('features')
                print(features[0:5])
                print()
                print('feature docs')
                print(feature_docs[0:5])
                cosim_matrix = cosine_sim(feature_docs)
                print()
                print('cosine sim matrix')
                print(type(cosim_matrix), len(cosim_matrix))
                print()

                print(cosim_matrix.shape)

                graphArray = []
                for i in range(len(cosim_matrix)):
                    for j in range(i):
                        if cosim_matrix[i][j] != 0 and cosim_matrix[i][j] != 1:
                            graphArray.append(cosim_matrix[i][j])

                # Similarity Thresholds we decided: (>0.3, >0.5, >0.7)
                '''
                <class 'numpy.ndarray'> 1682

                [[1.         0.         0.         ... 0.         0.34941857 0.        ]
                 [0.         1.         0.53676706 ... 0.         0.         0.        ]
                 [0.         0.53676706 1.         ... 0.         0.         0.        ]
                 [0.18860189 0.38145435 0.         ... 0.24094937 0.5397592  0.45125862]
                 [0.         0.30700538 0.57195272 ... 0.19392295 0.         0.36318585]
                 [0.         0.         0.         ... 0.53394963 0.         1.        ]]
                '''

                # print and plot histogram of similarites)
                # plt.hist(graphArray, 10)
                # plt.show()

            else:
                print('Empty dictionary, read in some data!')
                print()

        elif file_io == 'CBR-FE' or file_io == 'cbr-fe':
            print()
            # determine the U-I matrix to use ..
            if len(prefs) > 0 and len(prefs) <= 10:  # critics
                print('critics')
                userID = input(
                    'Enter username (for critics) or return to quit: ')
                movie_title_to_id = movie_to_ID(movies)
                recs = get_FE_recommendations(
                    prefs, features, movie_title_to_id, movies, userID)
                print(recs)
            elif len(prefs) > 10:
                print('ml-100k')
                userID = input(
                    'Enter userid (for ml-100k) or return to quit: ')
                movie_title_to_id = movie_to_ID(movies)
                recs = get_FE_recommendations(
                    prefs, features, movie_title_to_id, movies, userID)
                print(recs)
            else:
                print('Empty dictionary, read in some data!')
                print()

        elif file_io == 'CBR-TF' or file_io == 'cbr-tf':
            print()
            # determine the U-I matrix to use ..
            if len(prefs) > 0 and len(prefs) <= 10:  # critics
                print('critics')
                userID = input(
                    'Enter username (for critics) or return to quit: ')
                get_TFIDF_recommendations(
                    prefs, cosim_matrix, user=userID, sim_threshold=SIM_THRESHOLDS[0], movies=movies)

            elif len(prefs) > 10:
                print('ml-100k')
                userID = input(
                    'Enter userid (for ml-100k) or return to quit: ')
                get_TFIDF_recommendations(
                    prefs, cosim_matrix, user=userID, sim_threshold=SIM_THRESHOLDS[0], movies=movies)

            else:
                print('Empty dictionary, read in some data!')
                print()

        elif file_io == 'SIM' or file_io == 'sim':
            print()
            if len(prefs) > 0:
                ready = False  # sub command in progress
                sub_cmd = input(
                    'RD(ead) distance or RP(ead) pearson or WD(rite) distance or WP(rite) pearson? ')
                try:
                    if sub_cmd == 'RD' or sub_cmd == 'rd':
                        # Load the dictionary back from the pickle file.
                        itemsim = pickle.load(
                            open("save_itemsim_distance.p", "rb"))
                        sim_method = 'sim_distance'

                    elif sub_cmd == 'RP' or sub_cmd == 'rp':
                        # Load the dictionary back from the pickle file.
                        itemsim = pickle.load(
                            open("save_itemsim_pearson.p", "rb"))
                        sim_method = 'sim_pearson'

                    elif sub_cmd == 'WD' or sub_cmd == 'wd':
                        # transpose the U-I matrix and calc item-item similarities matrix
                        itemsim = calculateSimilarItems(
                            prefs, similarity=sim_distance, sim_weighting=SIM_WEIGHTING, sim_threshold=0)
                        # Dump/save dictionary to a pickle file
                        pickle.dump(itemsim, open(
                            "save_itemsim_distance.p", "wb"))
                        sim_method = 'sim_distance'

                    elif sub_cmd == 'WP' or sub_cmd == 'wp':
                        # transpose the U-I matrix and calc item-item similarities matrix
                        itemsim = calculateSimilarItems(
                            prefs, similarity=sim_pearson, sim_weighting=SIM_WEIGHTING, sim_threshold=0)
                        # Dump/save dictionary to a pickle file
                        pickle.dump(itemsim, open(
                            "save_itemsim_pearson.p", "wb"))
                        sim_method = 'sim_pearson'

                    else:
                        print("Sim sub-command %s is invalid, try again" % sub_cmd)
                        continue

                    ready = True  # sub command completed successfully

                except Exception as ex:
                    print('Error!!', ex, '\nNeed to W(rite) a file before you can R(ead) it!'
                          ' Enter Sim(ilarity matrix) again and choose a Write command')
                    print()

        elif file_io == 'HBR' or file_io == 'hbr':
            print()
            # determine the U-I matrix to use
            if len(cosim_matrix) > 0:
                if len(itemsim) > 0:
                    if len(prefs) > 0 and len(prefs) <= 10:  # critics
                        userID = input(
                            'Enter username (for critics) or return to quit: ')
                        weighting = input(
                            'Use weighting of {}? (T or F): '.format(HYBRID_WEIGHTING))
                        if weighting == "F" or weighting == "f":
                            print("weighting not selected")
                            get_hybrid_recommendations(
                                prefs, cosim_matrix, userID, SIM_THRESHOLDS[0], movies, itemsim, False)
                        else:
                            print("weighting selected")
                            get_hybrid_recommendations(
                                prefs, cosim_matrix, userID, SIM_THRESHOLDS[0], movies, itemsim, True)

                    elif len(prefs) > 10:
                        print('ml-100k')
                        userID = input(
                            'Enter userid (for ml-100k) or return to quit: ')
                        weighting = input(
                            'Use weighting of {}? (T or F): '.format(HYBRID_WEIGHTING))
                        if weighting == "F" or weighting == "f":
                            print("weighting not selected")
                            get_hybrid_recommendations(
                                prefs, cosim_matrix, userID, SIM_THRESHOLDS[0], movies, itemsim, False)
                        else:
                            print("weighting selected")
                            get_hybrid_recommendations(
                                prefs, cosim_matrix, userID, SIM_THRESHOLDS[0], movies, itemsim, True)
                    else:
                        print('Empty dictionary, read in some data!')
                        print()
                else:
                    print("Oops! Read a (Sim)ilarity Matrix first")
            else:
                print("Oops! Run TF-IDF first")

        elif file_io == 'LCVSIM' or file_io == 'lcvsim':
            print()
            sub_cmd = input('Select Recommender (FE, TFIDF, HBR): ')
            try:
                thissim = []
                othersim = itemsim

                if sub_cmd == 'HBR' or sub_cmd == 'hbr':
                    thissim = cosim_matrix
                    algo = get_hybrid_recommendations_single
                elif sub_cmd == 'FE' or sub_cmd == 'fe':
                    thissim = features
                    algo = get_FE_recommendations_single
                elif sub_cmd == 'TFIDF' or sub_cmd == 'tfidf':
                    thissim = cosim_matrix
                    algo = get_TFIDF_recommendations_single
                else:
                    print('Incorrect Command!')

                if len(prefs) > 0 and len(thissim) > 0:
                    # ask for hybrid weighting
                    if sub_cmd == 'HBR' or sub_cmd == 'hbr':
                        weighting = input(
                            'Use weighting of {}? (T or F): '.format(HYBRID_WEIGHTING))
                        if weighting == "T" or weighting == "t":
                            print("weighting selected")
                            weighting = True
                        else:
                            print("weighting not selected")
                            weighting = False

                    # run HBR/FE with sim_threshold > 0
                    if sub_cmd == 'HBR' or sub_cmd == 'hbr':
                        error_list = loo_cv_sim(
                            prefs, sim_method, algo, thissim, othersim, movies, SIM_THRESHOLDS[0], ws, row, weighting)
                        row += 1
                        print('len(SE list): %d, using %s' %
                              (len(error_list), sim_method))
                        print()
                        wb.save(dest)
                    elif sub_cmd == 'FE' or sub_cmd == 'fe':
                        error_list = loo_cv_sim(
                            prefs, "", algo, thissim, othersim, movies, SIM_THRESHOLDS[0], ws, row)
                        row += 1
                        print('len(SE list): %d' % (len(error_list)))
                        print()
                        wb.save(dest)
                    elif sub_cmd == 'TFIDF' or sub_cmd == 'tfidf':
                        # run TFIDF with thresholds t in SIM_THRESHOLDS = [0, 0.3, 0.5, 0.7]
                        print(
                            "TFIDF recognized. Running LCVSIM with all four similarity thresholds...")
                        for t in SIM_THRESHOLDS:
                            print("Running sim_threshold = {}".format(t))
                            error_list = loo_cv_sim(
                                prefs, "", algo, thissim, othersim, movies, t, ws, row)
                            row += 1
                            print('len(SE list): %d' % (len(error_list)))
                            print()
                            wb.save(dest)
                    else:
                        print('Incorrect Command!')
                else:
                    print('Empty dictionary, run R(ead) OR Empty Sim Matrix, run Sim!')

            except Exception as ex:
                print('Error!!', ex, '\nNeed to W(rite) a file before you can R(ead) it!'
                      ' Enter Sim(ilarity matrix) again and choose a Write command')
                print()
        else:
            done = True

    print('Goodbye!')


if __name__ == "__main__":
    main()


'''
Sample output ..


==>> cbr-fe
ml-100k

Enter username (for critics) or userid (for ml-100k) or return to quit: 340
rec for 340 = [
(5.0, 'Woman in Question, The (1950)'),
(5.0, 'Wallace & Gromit: The Best of Aardman Animation (1996)'),
(5.0, 'Thin Man, The (1934)'),
(5.0, 'Maltese Falcon, The (1941)'),
(5.0, 'Lost Highway (1997)'),
(5.0, 'Faust (1994)'),
(5.0, 'Daytrippers, The (1996)'),
(5.0, 'Big Sleep, The (1946)'),
(4.836990595611285, 'Sword in the Stone, The (1963)'),
(4.836990595611285, 'Swan Princess, The (1994)')]

==>> cbr-tf
ml-100k

Enter username (for critics) or userid (for ml-100k) or return to quit: 340
rec for 340 =  [
(5.000000000000001, 'Wallace & Gromit: The Best of Aardman Animation (1996)'),
(5.000000000000001, 'Faust (1994)'),
(5.0, 'Woman in Question, The (1950)'),
(5.0, 'Thin Man, The (1934)'),
(5.0, 'Maltese Falcon, The (1941)'),
(5.0, 'Lost Highway (1997)'),
(5.0, 'Daytrippers, The (1996)'),
(5.0, 'Big Sleep, The (1946)'),
(4.823001861184155, 'Sword in the Stone, The (1963)'),
(4.823001861184155, 'Swan Princess, The (1994)')]

'''
